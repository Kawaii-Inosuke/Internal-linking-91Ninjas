"""Ingestion pipeline (TRD §4/§5).

Reads a blog corpus (``.xlsx``/``.csv`` with columns ``link, title, content``),
chunks each article, embeds the chunks with a local model (``TASK_DOCUMENT``),
and stores pages + chunks in Postgres for a client.

Flow: read rows -> chunk (pure) -> embed everything in batches -> write pages +
chunks in one transaction. Chunking/embedding happen before any DB write, so a
failed run leaves the database untouched; a successful re-run refreshes each page
(delete-then-insert its chunks) rather than duplicating it.
"""
from __future__ import annotations

import csv
import logging
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from . import db
from .chunking import chunk_content, embedding_input
from .config import Config, ConfigError
from .embeddings import TASK_DOCUMENT, LocalEmbedder

log = logging.getLogger(__name__)

REQUIRED_COLUMNS = ("link", "title", "content")


@dataclass
class IngestStats:
    """Summary of an ingest run (for logging + M1 acceptance check)."""

    pages: int = 0
    chunks: int = 0
    skipped_rows: int = 0  # anomalies: content/title present but no link to key on
    blank_rows: int = 0  # fully empty spreadsheet rows (e.g. trailing blanks)
    zero_chunk_pages: list[str] = field(default_factory=list)


@dataclass
class _Row:
    url: str
    title: str
    content: str


@dataclass
class _PendingChunk:
    content: str  # stored content (title NOT included)
    embed_input: str  # title-prefixed text sent to the embedder


@dataclass
class _PendingPage:
    url: str
    title: str
    char_count: int
    chunks: list[_PendingChunk]


# --------------------------------------------------------------------------- #
# Reading the corpus
# --------------------------------------------------------------------------- #
def _cell(raw: tuple, index: int) -> str:
    """Return a stripped string for cell ``index`` of ``raw`` (empty if missing/None)."""
    if index >= len(raw):
        return ""
    value = raw[index]
    return "" if value is None else str(value).strip()


def _column_index(header: tuple) -> dict[str, int]:
    """Map each required column name to its position (case-insensitive)."""
    lookup: dict[str, int] = {}
    for i, name in enumerate(header):
        if name is None:
            continue
        lookup.setdefault(str(name).strip().lower(), i)
    missing = [col for col in REQUIRED_COLUMNS if col not in lookup]
    if missing:
        raise ValueError(
            f"Missing required column(s) {missing}; found columns {sorted(lookup)}."
        )
    return {col: lookup[col] for col in REQUIRED_COLUMNS}


def _read_xlsx(path: Path) -> list[_Row]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return []
        index = _column_index(header)
        return [
            _Row(
                url=_cell(raw, index["link"]),
                title=_cell(raw, index["title"]),
                content=_cell(raw, index["content"]),
            )
            for raw in rows_iter
        ]
    finally:
        workbook.close()


def _read_csv(path: Path) -> list[_Row]:
    # utf-8-sig strips a leading BOM (Excel "CSV UTF-8" export) so the first
    # header is "link", not "﻿link", which would fail column detection.
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            return []
        header = tuple(reader.fieldnames)
        index = _column_index(header)
        rows: list[_Row] = []
        for record in reader:
            values = tuple(record.get(name) for name in header)
            rows.append(
                _Row(
                    url=_cell(values, index["link"]),
                    title=_cell(values, index["title"]),
                    content=_cell(values, index["content"]),
                )
            )
        return rows


def _read_rows(path: Path) -> list[_Row]:
    """Read ``(link, title, content)`` rows from an ``.xlsx`` or ``.csv`` file."""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix == ".csv":
        return _read_csv(path)
    raise ValueError(f"Unsupported file type {suffix!r} (expected .xlsx or .csv).")


# --------------------------------------------------------------------------- #
# Ingestion
# --------------------------------------------------------------------------- #
def _plan_pages(rows: list[_Row], stats: IngestStats) -> list[_PendingPage]:
    """Chunk every row (no DB, no network) and record skipped/zero-chunk rows."""
    pending: list[_PendingPage] = []
    for row in rows:
        # Drop content-less rows BEFORE chunking/embedding: the corpus has ~835
        # blank trailing spreadsheet rows with no real content. Only content-
        # bearing posts should reach the embedder (M1 fix). Whitespace-only
        # content counts as blank.
        if not row.content.strip():
            stats.blank_rows += 1
            continue
        if not row.url:
            # A page is keyed by its URL, so a row with content but no link
            # cannot be ingested — that is a real anomaly, not a blank row.
            stats.skipped_rows += 1
            log.warning("Skipping row with content but no link (title=%r)", row.title[:60])
            continue
        chunks = chunk_content(row.content)
        pending_chunks = [
            _PendingChunk(content=chunk.content, embed_input=embedding_input(row.title, chunk.content))
            for chunk in chunks
        ]
        if not pending_chunks:
            stats.zero_chunk_pages.append(row.url)
        pending.append(
            _PendingPage(
                url=row.url,
                title=row.title,
                char_count=len(row.content),
                chunks=pending_chunks,
            )
        )
    return pending


def ingest_file(
    config: Config,
    file_path: str | Path,
    client: str,
    *,
    verify_model: bool = True,
) -> IngestStats:
    """Ingest a blog corpus file for ``client`` and return an :class:`IngestStats`.

    Set ``verify_model=False`` to skip the one-off embedding smoke test (used in
    tests). Raises :class:`FileNotFoundError` if the corpus is missing and
    ``ValueError`` for a malformed/unsupported file.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Corpus file not found: {path}")

    rows = _read_rows(path)
    log.info("Read %d rows from %s", len(rows), path.name)

    stats = IngestStats()
    pending = _plan_pages(rows, stats)

    embed_inputs = [chunk.embed_input for page in pending for chunk in page.chunks]

    # Connect first so DB problems (unreachable URL, EMBED_DIM != schema's vector
    # dimension) fail fast before the model is loaded and the corpus embedded.
    # The connection then stays idle (no open transaction) during embedding and
    # is reused for the write.
    conn = db.connect(config.database_url)
    try:
        schema_dim = db.embedding_column_dim(conn)
        if schema_dim is not None and schema_dim != config.embed_dim:
            raise ConfigError(
                f"EMBED_DIM={config.embed_dim} does not match the chunks.embedding "
                f"column dimension vector({schema_dim}) in schema.sql. "
                f"Align EMBED_DIM with the schema before ingesting."
            )
        conn.rollback()  # end the read transaction opened during connect/validation

        embedder = LocalEmbedder(
            config.embed_model,
            config.embed_dim,
            batch_size=config.embed_batch_size,
        )
        if verify_model:
            observed = embedder.smoke_test()
            log.info("Embedding model %s OK (dim=%d)", config.embed_model, observed)

        vectors = embedder.embed(embed_inputs, TASK_DOCUMENT) if embed_inputs else []
        if len(vectors) != len(embed_inputs):  # defensive: order-preserving stitch below
            raise RuntimeError(
                f"Embedder returned {len(vectors)} vectors for {len(embed_inputs)} inputs."
            )

        vector_iter = iter(vectors)
        with conn:  # one transaction: commit on success, roll back on error
            client_id = db.upsert_client(conn, client)
            for page in pending:
                page_id = db.upsert_page(
                    conn, client_id, page.url, page.title or None, page.char_count
                )
                db.delete_page_chunks(conn, page_id)
                stats.pages += 1
                chunk_rows = [
                    (idx, chunk.content, next(vector_iter))
                    for idx, chunk in enumerate(page.chunks)
                ]
                stats.chunks += db.insert_chunks(conn, client_id, page_id, chunk_rows)
    finally:
        conn.close()

    log.info(
        "Ingest complete: pages=%d chunks=%d skipped_rows=%d blank_rows=%d zero_chunk_pages=%d",
        stats.pages,
        stats.chunks,
        stats.skipped_rows,
        stats.blank_rows,
        len(stats.zero_chunk_pages),
    )
    if stats.zero_chunk_pages:
        log.warning("Pages that produced 0 chunks: %s", stats.zero_chunk_pages)
    return stats
